import openpyxl
import re

# Load the Excel file
workbook = openpyxl.load_workbook('/Users/adomoshagan/Desktop/Summer 23/GitHub Portfolio/Scheduling/JYCScheduling/data.xlsx')

cell_coordinates = {}
    
    
for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]
    sheet_coordinates = []
    
    # Iterate over each cell in the sheet
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value and re.match(r'\bavailable\b', str(cell.value), re.IGNORECASE):
                cell.value = sheet.title
                sheet_coordinates.append(cell.coordinate)
    cell_coordinates[sheet_name] = sheet_coordinates

# Save the modified workbook
workbook.save('modified_file.xlsx')
