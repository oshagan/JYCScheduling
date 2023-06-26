import openpyxl
import re
import random

# Load the Excel file
workbook = openpyxl.load_workbook('/Users/adomoshagan/Desktop/Summer 23/GitHub Portfolio/Scheduling/JYCScheduling/modified_file.xlsx')

cell_coordinates = {}

# Get the "Final" sheet
final_sheet = workbook['Final']

for sheet_name in workbook.sheetnames:
    if sheet_name == 'Final':
        continue  # Skip the "Final" sheet

    sheet = workbook[sheet_name]
    sheet_data = []

    # Iterate over each cell in the sheet
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == sheet.title:
                sheet_data.append((cell, cell.value))

    random.shuffle(sheet_data)  # Shuffle the list of cell data
    cell_coordinates[sheet_name] = sheet_data[:6]  # Take the first 6 random cell data

    # Place the cell data in the corresponding cells in the "Final" sheet
    for index, (cell, value) in enumerate(sheet_data[:6]):
        final_column = len(cell_coordinates) + 1
        
        # Check if the cell is part of a merged range
        for merged_range in sheet.merged_cells.ranges:
            if cell.coordinate in merged_range:
                # Unmerge the range
                sheet.unmerge_cells(merged_range.coord)
                # Write the value to each individual cell within the merged range
                for merged_cell in merged_range.cells:
                    final_cell = final_sheet.cell(row=merged_cell.row, column=final_column)
                    final_cell.value = value
            else:
                final_cell = final_sheet.cell(row=index + 1, column=final_column)
                final_cell.value = value

# Save the modified workbook
workbook.save('modified_file.xlsx')
