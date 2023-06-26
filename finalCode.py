#e2efda

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Load the existing Excel file
file_path = "/Users/adomoshagan/Desktop/Summer 23/GitHub Portfolio/Scheduling/JYCScheduling/data.xlsx"
workbook = load_workbook(filename=file_path)

# Iterate through each sheet in the workbook
for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]

    # Iterate through each cell in the sheet
    for row in sheet.iter_rows(min_row=4, max_row=29, min_col=3, max_col=9):
        for cell in row:
            if cell.value == "Available":
                cell.value = sheet_name

# Save the changes
workbook.save(file_path)

print("Cell text modification completed successfully.")
